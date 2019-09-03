'''
    ფუნქციები Excel-ის(xlsx) ფაილებთან მუშაობის გასამარტივებლად
'''


def cell_is_empty(cell_value):
    '''
    ამოწმებს, არის თუ არა Excel-ის ფაილის უჯრა ცარიელი
    არგუმენტები:
        1. cell_value - xlsx ფაილში უჯრის მნიშვნელობა
    '''
    if cell_value is None or not str(cell_value).strip():
        return True
    return False


def get_sheet_names(file, return_wb=False):
    '''
    გვიბრუნებს ფურცლების სიას კონკრეტული დოკუმენტიდან.
    თუ return_wb True-ა, დაბრუნებული შედეგის მე-2
    ელემენტი გახსნილი workbook ობიექტია.

    არგუმენტები:
        1. file - xlsx ფაილის მისამართი
        2. return_wb - დაგვიბრუნოს თუ არა workbook ობიექტი
                (ნაგულისხმევად=False). გამოსადეგია
                სისწრაფის თვალსაზრისით დიდი დოკუმენტებისთვის.
    '''
    import openpyxl

    wb = openpyxl.load_workbook(file)

    return wb.sheetnames if not return_wb else [wb.sheetnames, wb]


def get_last_row_num(file_or_wb_obj, sheet_name, column,
                     number=10, start_row=1):
    '''
        Excel-ის(xlsx) ფაილიდან ან შესაბამისი ობიექტიდან
    (openpyxl.workbook.workbook.Workbook)
    გვიბრუნებს ჩვენთვის სასურველი ფურცლის(sheet) სასურველი სვეტის(column)
    ბოლო სტრიქონის ნომერს.

        ბოლოდ მიიჩნევა ის სტრიქონი, რომლის ქვემოთაც გვხვდება მინიმუმ
    განსაზღვრული რაოდენობის(number) ცარიელი უჯრები.

    არგუმენტები:
        1. file_or_wb_obj  -  xlsx ფაილის მისამართი, ან ფაილის ობიექტი
                            (openpyxl.workbook.workbook.Workbook)

        2. sheet_name      -  ფურცლის(sheet) სახელი

        3. column          -  რომელი სვეტის გამოყენება გვინდა
                            ბოლო სტრიქონის საპოვნელად

        4. number          -  (ნაგულისხმევად = 10),
                            რამდენი ქვედა ცარიელი უჯრა ჩავთვალოთ
                            საკმარისად, რათა უჯრა ბოლოდ მივიჩნიოთ

        5. start_row       -  (ნაგულისხმევად = 1),
                            საწყისი სტრიქონი(შესაძლოა მონაცემები არ
                            იწყებოდეს პირველივე სტრიქონიდან)
    '''
    import openpyxl
    # load
    if not isinstance(file_or_wb_obj, openpyxl.workbook.workbook.Workbook):
        wb = openpyxl.load_workbook(file_or_wb_obj)
    else:
        wb = file_or_wb_obj

    # worksheet
    ws = wb[sheet_name]

    # to make code shorter
    r = start_row
    c = column

    # empty_num = 0
    empties_num = sum(
        [cell_is_empty(ws[f'{c}{r + i}'].value)
         for i in range(1, number + 1)])
    # if first is empty and there in next rows also
    # no value found, return 0
    if empties_num == number and cell_is_empty(ws[f'{c}{r}'].value):
        return 0
    else:
        # count better
        while empties_num != number:
            r += 1
            empties_num = sum(
                    [cell_is_empty(ws[f'{c}{r + i}'].value)
                     for i in range(1, number + 1)])
    return r


def get_data(
            file_or_wb_obj,
            check_column,
            start_row,
            data_columns,
            sheet_index=0,
            data_only=False,
            number=10,
            unpack_if_one=True,
            drop_empties=False):
    '''
        ფუნქცია გვეხმარება ჩვენთვის სასურველი სვეტების მონაცემების
    მიღებაში Excel-ის(xlsx) ფაილებიდან.

    #
        უცნაური მიზეზის გამო, თუ ფაილში ცვლილება ახლა შეგვაქვს
    და გვინდა განახლებული ფაილიდან ინფორმაციის მიღება, სასურველია
    ჯერ დავხუროთ Calc-ი, რადგან ზოგჯერ, სანამ პროგრამა
    გახსნილია, ფაილში ცვლილებები არ ჩანს Python-ისთვის.
        Excel-ის შემთხვევა ჯერ არ არის დატესტილი.
    #

        ჩვენ ვუთითებთ საწყის სტრიქონს და სვეტებს, რომლებიდანაც
    გვინდა მონაცემების მიღება.

        მონაცემები ბრუნდება list ტიპად, ქვე-list-ებით თითოეული
    სტრიქონისთვის, სვეტების იმ მიმდევრობით, რა მიმდევრობითაც მივუთითეთ
    data_columns არგუმენტში.

    არგუმენტები:
        1. file_or_wb_obj  -  Excel-ის ფაილის მისამართი,
                ან შესაბამისი ობიექტი(openpyxl.workbook.workbook.Workbook)

        2. check_column  -  სვეტი, რომლის გამოყენებაც გვინდა ბოლო
                            სტრიქონის იდენტიფიკაციისთვის
                            (სვეტი, რომელშიც ყველაზე მეტია ჩანაწერის
                             არსებობის ალბათობა)

        3. start_row  -  საწყისი სტრიქონის ნომერი,
                            საიდანაც გვინდა მონაცემების მიღება

        4. data_columns -  სვეტები, რომლებიდანაც გვინდა ინფორმაციის მიღება.
                        მაგალითი 1 სიმბოლოიანი სვეტების სახელებისთვის:
                            "ABCD"
                        მაგალითი 2 ან მეტ სიმბოლოიანი სვეტების სახელებისთვის:
                            ["AA", "BB", "AZ"]

        5. sheet_index -  (ნაგულისხმევად = 0, ანუ პირველი ფურცელი)

        6. data_only  -  (ნაგულისხმევად = False) - გამოვიყენოთ True,
                            თუ გვინდა xlsx ფაილში ფორმულების არსებობის
                            შემთხვევაში მათი მნიშვნელობები მივიღოთ
                            ფორმულების ნაცვლად.

        7. number  -  (ნაგულისხმევად = 10), რამდენი ქვედა ცარიელი უჯრა
                    ჩავთვალოთ საკმარისად, რათა უჯრა ბოლოდ მივიჩნიოთ

        8. unpack_if_one - (ნაგულისხმევად=True), თუ True-ა, როცა სვეტების
                        რაოდენობა, საიდანაც ვიღებთ მონაცემებს მხოლოდ ერთია,
                        შედეგი არის არა სიების სია, არამედ ერთი სია მხოლოდ
                        ამ სვეტის მნიშვნელობებით.

        9. drop_empties: - (ნაგულისხმევად=False), თუ True-ა, შედეგებში მივიღებთ
                        მხოლოდ არაცარიელ უჯრებს
                        (ცარიელად ითვლება მხოლოდ ცარიელი სივრცეც)
    '''
    import openpyxl
    # load excel data if workbook object is not directly used
    if not isinstance(file_or_wb_obj, openpyxl.workbook.workbook.Workbook):
        wb = openpyxl.load_workbook(
            file_or_wb_obj, data_only=data_only)
    else:
        wb = file_or_wb_obj

    # Get worksheet
    ws = wb.worksheets[sheet_index]

    # get number of rows we need
    last_row = get_last_row_num(file_or_wb_obj,
                                wb.worksheets[sheet_index].title,
                                check_column,
                                number=number,
                                start_row=start_row)
    # get actual data
    result_list = []

    for row in range(start_row, last_row+1):
        row_records = []

        for column in data_columns:
            row_records.append(ws[f'{column}{row}'].value)

        # remove empties if we wanted
        if drop_empties:
            if not any([not cell_is_empty(i) for i in row_records]):
                continue
        result_list.append(row_records)

    # unpack if needed
    if len(data_columns) == 1 and unpack_if_one:
        result_list = [i[0] for i in result_list]

    return result_list

def get_workbook_obj(xlsx_file):
    '''
    აბრუნებს openpyxl-ის workbook ობიექტს.
    შეცდომებზე შემოწმება არ ხდება.

    არგუმენტები:
        1. xlsx_file - xlsx ფაილის სახელი/მისამართი
    '''
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_file)

    return wb


def get_all_excel_column_letters():
    '''
    აბრუნებს xlsx ფაილის ყველა შესაძლო სვეტის ასოთა კომბინაციას
    '''
    import string

    uppers = string.ascii_uppercase
    # add 1 letters
    letters = list(uppers)
    # add 2 letters
    letters.extend([i + j for i in uppers for j in uppers])
    # add 3 letters
    letters.extend([i + j + k for i in uppers for j in uppers for k in uppers])

    return letters[:2**14]


def save_data(data,
              filename="New File.xlsx",
              sheet="Sheet 1",
              columns=None,
              start_row=1,
              bold_headers=True,
              center_data=True,
              ):
    '''
    # უკეთ დასატესტია! #

    ინახავს მონაცემებს xlsx ფაილში მითითებული სახელით,
    თუ ფაილი არსებობს, ეს ფაილი შეიცვლება,
    სხვა შემთხვევაში, იქმნება ახალი ფაილი.

    არგუმენტები:
        1. data - თუ ვწერთ ერთი სვეტის მონაცემებს list-ი,
                რომელიც შეიცავს ელემენტებს,
                    მაგ: [1, 2, 3, 4, 5]

                თუ ვწერთ ერთზე მეტი სვეტის მონაცემებს,
                list-ი, sublist-ებით, სადაც თითოეული sublist-ის ელემენტი
                სწორი თანმიმდევრობით შეესაბამება სვეტების სახელებს, რომელსაც
                მივუთითებთ columns არგუმენტში.
                    მაგ: [
                        [  header_1,     header_2,    header_3 ],
                        [ row_1_col_1,  row_1_col_2, row_1_col_3]
                        [ row_2_col_1,  row_2_col_2, row_2_col_3]
                        [ row_3_col_1,  row_3_col_2, row_3_col_3]
                    ]

        2. filename - ფაილის სახელი/მისამართი.
                    თუ ფაილი არსებობს, ცვლილებები შევა მასში, თუ
                    არ არსებობს, შეიქმნება და შეინახება მითითებული სახელით

        3. sheet - ფურცლის სახელი. თუ არ არსებობს, შეიქმნება
                    (ნაგულისხმევად="Sheet 1").

        4. columns - სვეტები, რომლებშიც გვინდა მონაცემთა შენახვა.
                    თუ სვეტების სახელები მხოლოდ ერთ ასოიანია, შესაძლებელია
                    ტექსტური ტიპის გამოყენება (მაგ: "ABC").

                    სხვა შემთხვევაში, საჭიროა ეს მონაცემები იყოს სიაში.
                    (მაგ: ["PY", "TH", "ON"])

        5. start_row - რომელი სტრიქონიდან დაიწყოს მონაცემთა ჩაწერა
                        (ნაგულისხმევად=1)

        6. bold_headers - გავამუქოთ თუ არა headers(პირველი ელემენტი data-ში)
                        (ნაგულისხმევად=True)

        7. center_data - მოვათავსოთ თუ არა ყველა მონაცემი ცენტრში
                        (ნაგულისხმევად=True)
    '''

    import os
    import openpyxl
    from openpyxl.styles import Font, Alignment

    if not isinstance(data, list):
        raise Exception(
                f"Please use list for data argument, not {type(data)}")

    xlsx_file = filename + ".xlsx"

    # convert [1,2,3] to [[1], [2], [3]] to make things easier later
    if not isinstance(data[0], list):
        data = [[i] for i in data]

    # get workbook
    if os.path.isfile(xlsx_file):
        wb = openpyxl.load_workbook(xlsx_file)
        print("Already Created File loaded")
    else:
        # delete default sheet
        wb = openpyxl.Workbook()
        wb.remove(wb["Sheet"])

    # get column names
    if columns is not None:
        assert isinstance(columns, (list, str))

        if len(columns) != len(data[0]):
            raise Exception("Number of columns in first row is "
                            "not same as given columns as argument"
                            f"({len(data[0])}!={len(columns)})")
    else:
        columns = get_all_excel_column_letters()[:len(data[0])]

    # get worksheet
    if sheet in wb.sheetnames:
        ws = wb[sheet]
    else:
        ws = wb.create_sheet()
        ws.title = sheet

    # write data in file
    for index, data_in_row in enumerate(data):
        for data_in_cell, column in zip(data_in_row, columns):
            ws[f'{column}{index + start_row}'].value = data_in_cell

    # change header widths
    for index, column in enumerate(columns):
        ws.column_dimensions[column].width = \
                                max(int(len(str(data[0][index]))*1.2) + 1, 5)

    # make headers bold
    if bold_headers:
        for i in ws[f"{start_row}:{start_row}"]:
            i.font = Font(bold=True)

    # center rows if necessary
    if center_data:
        for index, data_in_row in enumerate(data):
            for data_in_cell, column in zip(data_in_row, columns):
                ws[f'{column}{index + start_row}'].alignment = \
                                                Alignment(horizontal='center')

    wb.save(xlsx_file)


def xls_to_xlsx(xls_file_path):
    '''
       აკონვერტირებს xls ფაილს xlsx ფაილად და ინახავს
       იმავე სახელით, ახალი გაფართოებით.

       #! გამოყენებამდე აუცილებელია
            pyexcel, pyexcel-xls და pyexcel-xlsxw ბიბლიოთეკების ინსტალაცია !#

       არგუმენტები:
            1. xls_file_path - xls ფაილის მისამართი
    '''
    # replace xls files with xlsx files
    import pyexcel

    if not xls_file_path.lower().endswith(".xls"):
        raise Exception(
                    "Please use xls file, not ",
                    xls_file_path.split(".")[-1])

    pyexcel.save_book_as(file_name=xls_file_path,
                         dest_file_name=xls_file_path + "x")
